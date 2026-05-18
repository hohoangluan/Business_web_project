"""
Views cho stats_reports — thống kê tổng hợp, export CSV, print PDF.
"""

import csv
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.http import HttpResponse

from accounts.services import ensure_profile, can_access_statistics
from stats_reports.services import (
    build_statistics_page_context,
    STATISTICS_TYPE_LABEL_MAP,
)


@login_required
def statistics_view(request):
    """Trang statistics. Template: stats_reports/statistics.html"""
    ensure_profile(request.user)
    if not can_access_statistics(request.user):
        messages.error(request, 'Bạn không có quyền xem statistics.')
        return redirect('dashboard')

    context = build_statistics_page_context(request.user, request.GET)
    context['active_page'] = 'statistics'
    return render(request, 'stats_reports/statistics.html', context)


import openpyxl
from openpyxl.styles import Font, Alignment

@login_required
def statistics_export_excel_view(request):
    """Xuất bảng tổng hợp statistics ra file Excel (.xlsx)."""
    ensure_profile(request.user)
    if not can_access_statistics(request.user):
        messages.error(request, 'Bạn không có quyền xuất statistics.')
        return redirect('dashboard')

    ctx = build_statistics_page_context(request.user, request.GET)
    tr = ctx['time_range']
    st = ctx['selected_stats_type']
    fname = f"statistics_{st}_{tr['start_date']:%Y%m%d}_{tr['end_date']:%Y%m%d}.xlsx"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo Thống kê"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    
    ws.append(['Báo cáo statistics'])
    ws['A1'].font = title_font
    
    ws.append(['Loại thống kê', ctx['selected_stats_type_label']])
    for item in ctx['statistics_sections']['filter_summary']:
        ws.append([item])
    ws.append([])

    sections = ctx['statistics_sections']

    def append_header(headers):
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font

    if st == 'evaluation':
        append_header(['Nhân viên', 'Username', 'Phòng ban', 'Người đánh giá', 'Ngày', 'Nội dung', 'Minh chứng'])
        for r in sections['evaluation_rows']:
            ws.append([r['employee_name'], r['employee_username'], r['department'], f"{r['reviewer_name']} ({r['reviewer_role']})", r['evaluation_date_display'], r['evaluation_content'], r['evidence_reference']])
        wb.save(response)
        return response

    if st == 'rewards':
        append_header(['Nhân viên', 'Username', 'Phòng ban', 'Người đề xuất', 'Phân loại', 'Số tiền', 'Ngày', 'Trạng thái', 'Lý do'])
        for r in sections['rewards_rows']:
            ws.append([r['employee_name'], r['employee_username'], r['department'], f"{r['proposer_name']} ({r['proposer_role']})", r['type_label'], r['amount_display'], r['application_date_display'], r['status'], r['reason_title']])
        wb.save(response)
        return response

    if st == 'leave':
        append_header(['Nhân viên', 'Username', 'Phòng ban', 'Quản lý', 'Leader', 'Ngày nghỉ', 'Số đơn'])
        for r in sections['summary_rows']:
            ws.append([r['employee_name'], r['employee_username'], r['department'], r['manager_name'], r['leader_name'], r['leave_days'], r['leave_requests']])
        wb.save(response)
        return response

    if st == 'attendance':
        append_header(['Nhân viên', 'Username', 'Phòng ban', 'Giờ tăng ca', 'Đi trễ', 'Nghỉ làm', 'Tỷ lệ'])
        for r in sections['summary_rows']:
            ws.append([r['employee_name'], r['employee_username'], r['department'], r['overtime_hours'], r['late_count'], r['absence_days'], r['attendance_rate']])
        wb.save(response)
        return response

    append_header(['Nhân viên', 'Username', 'Phòng ban', 'Quản lý', 'Leader', 'Ngày nghỉ', 'Số đơn', 'Giờ tăng ca', 'Đi trễ', 'Nghỉ làm', 'Tỷ lệ'])
    for r in sections['summary_rows']:
        ws.append([r['employee_name'], r['employee_username'], r['department'], r['manager_name'], r['leader_name'], r['leave_days'], r['leave_requests'], r['overtime_hours'], r['late_count'], r['absence_days'], r['attendance_rate']])

    if st == 'all':
        ws.append([])
        ws.append(['Thống kê đánh giá'])
        ws['A' + str(ws.max_row)].font = title_font
        append_header(['Nhân viên', 'Username', 'Phòng ban', 'Người đánh giá', 'Ngày', 'Nội dung', 'Minh chứng'])
        for r in sections['evaluation_rows']:
            ws.append([r['employee_name'], r['employee_username'], r['department'], f"{r['reviewer_name']} ({r['reviewer_role']})", r['evaluation_date_display'], r['evaluation_content'], r['evidence_reference']])
        
        ws.append([])
        ws.append(['Thống kê khen thưởng và xử phạt'])
        ws['A' + str(ws.max_row)].font = title_font
        append_header(['Nhân viên', 'Username', 'Phòng ban', 'Người đề xuất', 'Phân loại', 'Số tiền', 'Ngày', 'Trạng thái', 'Lý do'])
        for r in sections['rewards_rows']:
            ws.append([r['employee_name'], r['employee_username'], r['department'], f"{r['proposer_name']} ({r['proposer_role']})", r['type_label'], r['amount_display'], r['application_date_display'], r['status'], r['reason_title']])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


@login_required
def statistics_print_view(request):
    """Trang in tối giản. Template: stats_reports/statistics_print.html"""
    ensure_profile(request.user)
    if not can_access_statistics(request.user):
        messages.error(request, 'Bạn không có quyền in statistics.')
        return redirect('dashboard')

    context = build_statistics_page_context(request.user, request.GET)
    return render(request, 'stats_reports/statistics_print.html', context)
